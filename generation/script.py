# from openpyxl import load_workbook
# from openpyxl import Workbook
from connect import execute_query
from connect import connection
import openpyxl as op
import json


# ['Титул лист', 'Сводные', 'график учебного процесса', 'План', 'Расшифровка', 'Кабинеты', 'ПЗ ', 'Комплексный', 'Лист изменений', 'Приложение 1']
# stringuxa = '{"tema_1" : {0, 0}, "tema_2" : {0, 0}, "tema_3" : {0, 0}, "tema_4" : {0, 0}, "tema_5" : {0, 0}, "tema_6" : {0, 0}, "tema_7" : {0, 0}, "tema_8" : {0, 0}, "tema_9" : {0, 0}, "tema_10" : {0, 0}, "tema_11" : {0, 0}, "tema_12" : {0, 0}, "tema_13" : {0, 0}, "tema_14" : {0, 0}, "tema_15" : {0, 0}, "tema_16" : {0, 0}, "tema_17" : {0, 0}, "tema_18" : {0, 0}, "tema_19" : {0, 0}, "tema_20" : {0, 0}, "tema_21" : {0, 0}, "tema_22" : {0, 0}, "tema_23" : {0, 0}, "tema_24" : {0, 0}, "tema_25" : {0, 0}, "tema_26" : {0, 0}, "tema_27" : {0, 0}, "tema_28" : {0, 0}, "tema_29" : {0, 0}, "tema_30" : {0, 0}, "tema_31" : {0, 0}, "tema_32" : {0, 0}, "tema_33" : {0, 0}, "tema_34" : {0, 0}, "tema_35" : {0, 0}, "tema_36" : {0, 0}, "tema_37" : {0, 0}, "tema_38" : {0, 0}, "tema_39" : {0, 0}, "tema_40" : {0, 0}, "tema_41" : {0, 0}, "tema_42" : {0, 0}, "tema_43" : {0, 0}, "tema_44" : {0, 0}, "tema_45" : {0, 0}, "tema_46" : {0, 0}, "tema_47" : {0, 0}, "tema_48" : {0, 0}, "tema_49" : {0, 0}, "tema_50" : {0, 0}, "tema_51" : {0, 0}, "tema_52" : {0, 0}, "tema_53" : {0, 0}, "tema_54" : {0, 0}, "tema_55" : {0, 0}, "tema_56" : {0, 0}, "tema_57" : {0, 0}, "tema_58" : {0, 0}, "tema_59" : {0, 0}, "tema_60" : {0, 0}, "tema_61" : {0, 0}, "tema_62" : {0, 0}, "tema_63" : {0, 0}, "tema_64" : {0, 0}, "tema_65" : {0, 0}, "tema_66" : {0, 0}, "tema_67" : {0, 0}, "tema_68" : {0, 0}, "tema_69" : {0, 0}, "tema_70" : {0, 0}, "tema_71" : {0, 0}, "tema_72" : {0, 0}, "tema_73" : {0, 0}, "tema_74" : {0, 0}, "tema_75" : {0, 0}, "tema_76" : {0, 0}, "tema_77" : {0, 0}, "tema_78" : {0, 0}, "tema_79" : {0, 0}, "tema_80" : {0, 0}, "tema_81" : {0, 0}, "tema_82" : {0, 0}, "tema_83" : {0, 0}, "tema_84" : {0, 0}, "tema_85" : {0, 0}, "tema_86" : {0, 0}, "tema_87" : {0, 0}, "tema_88" : {0, 0}, "tema_89" : {0, 0}, "tema_90" : {0, 0}, "tema_91" : {0, 0}, "tema_92" : {0, 0}, "tema_93" : {0, 0}, "tema_94" : {0, 0}, "tema_95" : {0, 0}, "tema_96" : {0, 0}, "tema_97" : {0, 0}, "tema_98" : {0, 0}, "tema_99" : {0, 0}, "tema_100" : {0, 0}, "tema_101" : {0, 0}, "tema_102" : {0, 0}, "tema_103" : {0, 0}, "tema_104" : {0, 0}, "tema_105" : {0, 0}, "tema_106" : {0, 0}, "tema_107" : {0, 0}, "tema_108" : {0, 0}, "tema_109" : {0, 0}, "tema_110" : {0, 0}, "tema_111" : {0, 0}, "tema_112" : {0, 0}, "tema_113" : {0, 0}, "tema_114" : {0, 0}, "tema_115" : {0, 0}, "tema_116" : {0, 0}, "tema_117" : {0, 0}, "tema_118" : {0, 0}, "tema_119" : {0, 0}, "tema_120" : {0, 0}, "tema_121" : {0, 0}, "tema_122" : {0, 0}, "tema_123" : {0, 0}, "tema_124" : {0, 0}, "tema_125" : {0, 0}, "tema_126" : {0, 0}, "tema_127" : {0, 0}, "tema_128" : {0, 0}, "tema_129" : {0, 0}, "tema_130" : {0, 0}, "tema_131" : {0, 0}, "tema_132" : {0, 0}, "tema_133" : {0, 0}, "tema_134" : {0, 0}, "tema_135" : {0, 0}, "tema_136" : {0, 0}, "tema_137" : {0, 0}, "tema_138" : {0, 0}, "tema_139" : {0, 0}, "tema_140" : {0, 0}, "tema_141" : {0, 0}, "tema_142" : {0, 0}, "tema_143" : {0, 0}, "tema_144" : {0, 0}, "tema_145" : {0, 0}, "tema_146" : {0, 0}, "tema_147" : {0, 0}, "tema_148" : {0, 0}, "tema_149" : {0, 0}, "tema_150" : {0, 0}}'
stringuxa = '{"tema_1" : {"complete" : 0, "date": 0}, "tema_2" : {"complete" : 0, "date": 0}, "tema_3" : {"complete" : 0, "date": 0}, "tema_4" : {"complete" : 0, "date": 0}, "tema_5" : {"complete" : 0, "date": 0}, "tema_6" : {"complete" : 0, "date": 0}, "tema_7" : {"complete" : 0, "date": 0}, "tema_8" : {"complete" : 0, "date": 0}, "tema_9" : {"complete" : 0, "date": 0}, "tema_10" : {"complete" : 0, "date": 0}, "tema_11" : {"complete" : 0, "date": 0}, "tema_12" : {"complete" : 0, "date": 0}, "tema_13" : {"complete" : 0, "date": 0}, "tema_14" : {"complete" : 0, "date": 0}, "tema_15" : {"complete" : 0, "date": 0}, "tema_16" : {"complete" : 0, "date": 0}, "tema_17" : {"complete" : 0, "date": 0}, "tema_18" : {"complete" : 0, "date": 0}, "tema_19" : {"complete" : 0, "date": 0}, "tema_20" : {"complete" : 0, "date": 0}, "tema_21" : {"complete" : 0, "date": 0}, "tema_22" : {"complete" : 0, "date": 0}, "tema_23" : {"complete" : 0, "date": 0}, "tema_24" : {"complete" : 0, "date": 0}, "tema_25" : {"complete" : 0, "date": 0}, "tema_26" : {"complete" : 0, "date": 0}, "tema_27" : {"complete" : 0, "date": 0}, "tema_28" : {"complete" : 0, "date": 0}, "tema_29" : {"complete" : 0, "date": 0}, "tema_30" : {"complete" : 0, "date": 0}, "tema_31" : {"complete" : 0, "date": 0}, "tema_32" : {"complete" : 0, "date": 0}, "tema_33" : {"complete" : 0, "date": 0}, "tema_34" : {"complete" : 0, "date": 0}, "tema_35" : {"complete" : 0, "date": 0}, "tema_36" : {"complete" : 0, "date": 0}, "tema_37" : {"complete" : 0, "date": 0}, "tema_38" : {"complete" : 0, "date": 0}, "tema_39" : {"complete" : 0, "date": 0}, "tema_40" : {"complete" : 0, "date": 0}, "tema_41" : {"complete" : 0, "date": 0}, "tema_42" : {"complete" : 0, "date": 0}, "tema_43" : {"complete" : 0, "date": 0}, "tema_44" : {"complete" : 0, "date": 0}, "tema_45" : {"complete" : 0, "date": 0}, "tema_46" : {"complete" : 0, "date": 0}, "tema_47" : {"complete" : 0, "date": 0}, "tema_48" : {"complete" : 0, "date": 0}, "tema_49" : {"complete" : 0, "date": 0}, "tema_50" : {"complete" : 0, "date": 0}, "tema_51" : {"complete" : 0, "date": 0}, "tema_52" : {"complete" : 0, "date": 0}, "tema_53" : {"complete" : 0, "date": 0}, "tema_54" : {"complete" : 0, "date": 0}, "tema_55" : {"complete" : 0, "date": 0}, "tema_56" : {"complete" : 0, "date": 0}, "tema_57" : {"complete" : 0, "date": 0}, "tema_58" : {"complete" : 0, "date": 0}, "tema_59" : {"complete" : 0, "date": 0}, "tema_60" : {"complete" : 0, "date": 0}, "tema_61" : {"complete" : 0, "date": 0}, "tema_62" : {"complete" : 0, "date": 0}, "tema_63" : {"complete" : 0, "date": 0}, "tema_64" : {"complete" : 0, "date": 0}, "tema_65" : {"complete" : 0, "date": 0}, "tema_66" : {"complete" : 0, "date": 0}, "tema_67" : {"complete" : 0, "date": 0}, "tema_68" : {"complete" : 0, "date": 0}, "tema_69" : {"complete" : 0, "date": 0}, "tema_70" : {"complete" : 0, "date": 0}, "tema_71" : {"complete" : 0, "date": 0}, "tema_72" : {"complete" : 0, "date": 0}, "tema_73" : {"complete" : 0, "date": 0}, "tema_74" : {"complete" : 0, "date": 0}, "tema_75" : {"complete" : 0, "date": 0}, "tema_76" : {"complete" : 0, "date": 0}, "tema_77" : {"complete" : 0, "date": 0}, "tema_78" : {"complete" : 0, "date": 0}, "tema_79" : {"complete" : 0, "date": 0}, "tema_80" : {"complete" : 0, "date": 0}, "tema_81" : {"complete" : 0, "date": 0}, "tema_82" : {"complete" : 0, "date": 0}, "tema_83" : {"complete" : 0, "date": 0}, "tema_84" : {"complete" : 0, "date": 0}, "tema_85" : {"complete" : 0, "date": 0}, "tema_86" : {"complete" : 0, "date": 0}, "tema_87" : {"complete" : 0, "date": 0}, "tema_88" : {"complete" : 0, "date": 0}, "tema_89" : {"complete" : 0, "date": 0}, "tema_90" : {"complete" : 0, "date": 0}, "tema_91" : {"complete" : 0, "date": 0}, "tema_92" : {"complete" : 0, "date": 0}, "tema_93" : {"complete" : 0, "date": 0}, "tema_94" : {"complete" : 0, "date": 0}, "tema_95" : {"complete" : 0, "date": 0}, "tema_96" : {"complete" : 0, "date": 0}, "tema_97" : {"complete" : 0, "date": 0}, "tema_98" : {"complete" : 0, "date": 0}, "tema_99" : {"complete" : 0, "date": 0}, "tema_100" : {"complete" : 0, "date": 0}, "tema_101" : {"complete" : 0, "date": 0}, "tema_102" : {"complete" : 0, "date": 0}, "tema_103" : {"complete" : 0, "date": 0}, "tema_104" : {"complete" : 0, "date": 0}, "tema_105" : {"complete" : 0, "date": 0}, "tema_106" : {"complete" : 0, "date": 0}, "tema_107" : {"complete" : 0, "date": 0}, "tema_108" : {"complete" : 0, "date": 0}, "tema_109" : {"complete" : 0, "date": 0}, "tema_110" : {"complete" : 0, "date": 0}, "tema_111" : {"complete" : 0, "date": 0}, "tema_112" : {"complete" : 0, "date": 0}, "tema_113" : {"complete" : 0, "date": 0}, "tema_114" : {"complete" : 0, "date": 0}, "tema_115" : {"complete" : 0, "date": 0}, "tema_116" : {"complete" : 0, "date": 0}, "tema_117" : {"complete" : 0, "date": 0}, "tema_118" : {"complete" : 0, "date": 0}, "tema_119" : {"complete" : 0, "date": 0}, "tema_120" : {"complete" : 0, "date": 0}, "tema_121" : {"complete" : 0, "date": 0}, "tema_122" : {"complete" : 0, "date": 0}, "tema_123" : {"complete" : 0, "date": 0}, "tema_124" : {"complete" : 0, "date": 0}, "tema_125" : {"complete" : 0, "date": 0}, "tema_126" : {"complete" : 0, "date": 0}, "tema_127" : {"complete" : 0, "date": 0}, "tema_128" : {"complete" : 0, "date": 0}, "tema_129" : {"complete" : 0, "date": 0}, "tema_130" : {"complete" : 0, "date": 0}, "tema_131" : {"complete" : 0, "date": 0}, "tema_132" : {"complete" : 0, "date": 0}, "tema_133" : {"complete" : 0, "date": 0}, "tema_134" : {"complete" : 0, "date": 0}, "tema_135" : {"complete" : 0, "date": 0}, "tema_136" : {"complete" : 0, "date": 0}, "tema_137" : {"complete" : 0, "date": 0}, "tema_138" : {"complete" : 0, "date": 0}, "tema_139" : {"complete" : 0, "date": 0}, "tema_140" : {"complete" : 0, "date": 0}, "tema_141" : {"complete" : 0, "date": 0}, "tema_142" : {"complete" : 0, "date": 0}, "tema_143" : {"complete" : 0, "date": 0}, "tema_144" : {"complete" : 0, "date": 0}, "tema_145" : {"complete" : 0, "date": 0}, "tema_146" : {"complete" : 0, "date": 0}, "tema_147" : {"complete" : 0, "date": 0}, "tema_148" : {"complete" : 0, "date": 0}, "tema_149" : {"complete" : 0, "date": 0}, "tema_150" : {"complete" : 0, "date": 0}}'
stringuxa = json.dumps(stringuxa)

filename = "excel.xlsx"
wb = op.load_workbook(filename, data_only=True)
sheet = wb.active
print(sheet.cell(row=8, column=2).value)
table_number = 20
for i in range(8, 62):
#НАЗВАНИЕ ПРЕДМЕТА
    # if sheet.cell(row=i, column=1).value != None:
    #     discName1 = sheet.cell(row=i, column=1).value
    #     discName2 = sheet.cell(row=i, column=2).value
    #     discName = str(discName1) + ' ' + str(discName2)
    # else:
    #     discName = sheet.cell(row=i, column=2).value

    discName = sheet.cell(row=i, column=2).value
# КОЛ-ВО ЧАСОВ ПО ПРЕДМЕТУ
    countHours = sheet.cell(row=i, column=7).value
# САМОСТОЯТЕЛЬНАЯ РАБОТА
    if sheet.cell(row=i, column=8).value != None:
        samostAttestation = sheet.cell(row=i, column=8).value
    else:
        egz = 0
# "ЭКЗАМЕНЫ"
    if sheet.cell(row=i, column=3).value != None:
        egz = sheet.cell(row=i, column=3).value
    else:
        egz = 0
# "ДИФ. ЗАЧЁТЫ"
    if sheet.cell(row=i, column=4).value != None:
        difZachet = sheet.cell(row=i, column=4).value
    else:
        difZachet = 0
# "ЗАЧЁТЫ"
    if sheet.cell(row=i, column=5).value != None:
        zachet = sheet.cell(row=i, column=5).value
    else:
        zachet = 0
# "ДРУГАЯ АТТЕСТАЦИЯ"
    if sheet.cell(row=i, column=6).value != None:
        anotherAttestation = sheet.cell(row=i, column=6).value
    else:
        anotherAttestation = 0
#ТЕОРЕТИЧЕСКИХ ЗАНЯТИЙ
    if sheet.cell(row=i, column=10).value != None:
        teorLession = sheet.cell(row=i, column=10).value
    else:
        teorLession = 0
#ЛАБОРАТОРНЫХ ЗАНЯТИЙ
    if sheet.cell(row=i, column=11).value != None:
        laborLession = sheet.cell(row=i, column=11).value
    else:
        laborLession = 0
#КУРСОВЫХ(ПРОЕКТОВ)
    if sheet.cell(row=i, column=12).value != None:
        courseProgect = sheet.cell(row=i, column=12).value
    else:
        courseProgect = 0
#КОЛ-ВО ПРАКТИКИ ПО ПРЕДМЕТУ
    if sheet.cell(row=i, column=13).value != None:
        practice = sheet.cell(row=i, column=13).value
    else:
        practice = 0
#КОНСУЛЬТАЦИИ
    if sheet.cell(row=i, column=14).value != None:
        consultation = sheet.cell(row=i, column=14).value
    else:
        consultation = 0
#ПРОМЕЖУТОЧНАЯ АТТЕСТАЦИЯ
    if sheet.cell(row=i, column=15).value != None:
        promezgAttest = sheet.cell(row=i, column=15).value
    else:
        promezgAttest = 0
# ЗАПОЛЕНИЕ ЧАСОВ ПО КАЖДОМУ СЕМЕСРУ
    if sheet.cell(row=i, column=16).value != None:
        semestr1 = sheet.cell(row=i, column=16).value
    else:
        semestr1 = 0
    if sheet.cell(row=i, column=17).value != None:
        semestr2 = sheet.cell(row=i, column=17).value
    else:
        semestr2 = 0
    if sheet.cell(row=i, column=18).value != None:
        semestr3 = sheet.cell(row=i, column=18).value
    else:
        semestr3 = 0
    if sheet.cell(row=i, column=19).value != None:
        semestr4 = sheet.cell(row=i, column=19).value
    else:
        semestr4 = 0
    if sheet.cell(row=i, column=20).value != None:
            semestr5 = sheet.cell(row=i, column=20).value
    else:
        semestr5 = 0
    if sheet.cell(row=i, column=21).value != None:
        semestr6 = sheet.cell(row=i, column=21).value
    else:
        semestr6 = 0
    if sheet.cell(row=i, column=22).value != None:
        semestr7 = sheet.cell(row=i, column=22).value
    else:
        semestr7 = 0
    if sheet.cell(row=i, column=23).value != None:
        semestr8 = sheet.cell(row=i, column=23).value
    else:
        semestr8 = 0

# print(discName, ' ', countHours, ' ', egz, ' ', difZachet, ' ', zachet, ' ', anotherAttestation, ' ', samostAttestation,
#       ' ',
#       ' Сколько в кажом семетре часов: ', semestr1, ' ', semestr2, ' ', semestr3, ' ', semestr4, ' ', semestr5, ' ',
#       semestr6, ' '
#       , semestr7, ' ', semestr8, ' ')
    sql = f"INSERT INTO `lessons_hours` (`discName`, `group_number`, `exam`, `diff_exam`, `zachet`, `other_attestation`, `total`," \
          " `samost_work`, `theoretical_lesson`, `laboratory_lesson`, `coursovie`, `practice`, `consultation`, `promezhutAttest`," \
          " `semestr1`, `semestr2`, `semestr3`, `semestr4`, `semestr5`, `semestr6`, `semestr7`, `semestr8`, `table_number`, `completed_state`) " \
          f"VALUES ('{discName}', '9КС-22', '{egz}', '{difZachet}', '{zachet}', '{anotherAttestation}', {countHours}," \
          f" {samostAttestation}, {teorLession}, {laborLession}, {courseProgect}, {practice}, {consultation}" \
          f", {promezgAttest}, {semestr1}, {semestr2}, {semestr3}, {semestr4}, {semestr5}, {semestr6}," \
          f" {semestr7}, {semestr8}, {table_number}, {stringuxa})"
    try:
        execute_query(connection, sql)
    except:
        print(discName)
    table_number += 1


# ALTER TABLE `disciplina_3` ADD `tema_51` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_50`;
# ALTER TABLE `disciplina_3` ADD `tema_52` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_51`;
# ALTER TABLE `disciplina_3` ADD `tema_53` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_52`;
# ALTER TABLE `disciplina_3` ADD `tema_54` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_53`;
# ALTER TABLE `disciplina_3` ADD `tema_55` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_54`;
# ALTER TABLE `disciplina_3` ADD `tema_56` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_55`;
# ALTER TABLE `disciplina_3` ADD `tema_57` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_56`;
# ALTER TABLE `disciplina_3` ADD `tema_58` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_57`;
# ALTER TABLE `disciplina_3` ADD `tema_59` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_58`;
# ALTER TABLE `disciplina_3` ADD `tema_60` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_59`;
# ALTER TABLE `disciplina_3` ADD `tema_61` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_60`;
# ALTER TABLE `disciplina_3` ADD `tema_62` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_61`;
# ALTER TABLE `disciplina_3` ADD `tema_63` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_62`;
# ALTER TABLE `disciplina_3` ADD `tema_64` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_63`;
# ALTER TABLE `disciplina_3` ADD `tema_65` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_64`;
# ALTER TABLE `disciplina_3` ADD `tema_66` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_65`;
# ALTER TABLE `disciplina_3` ADD `tema_67` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_66`;
# ALTER TABLE `disciplina_3` ADD `tema_68` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_67`;
# ALTER TABLE `disciplina_3` ADD `tema_69` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_68`;
# ALTER TABLE `disciplina_3` ADD `tema_70` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_69`;
# ALTER TABLE `disciplina_3` ADD `tema_71` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_70`;
# ALTER TABLE `disciplina_3` ADD `tema_72` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_71`;
# ALTER TABLE `disciplina_3` ADD `tema_73` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_72`;
# ALTER TABLE `disciplina_3` ADD `tema_74` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_73`;
# ALTER TABLE `disciplina_3` ADD `tema_75` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_74`;
# ALTER TABLE `disciplina_3` ADD `tema_76` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_75`;
# ALTER TABLE `disciplina_3` ADD `tema_77` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_76`;
# ALTER TABLE `disciplina_3` ADD `tema_78` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_77`;
# ALTER TABLE `disciplina_3` ADD `tema_79` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_78`;
# ALTER TABLE `disciplina_3` ADD `tema_80` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_79`;
# ALTER TABLE `disciplina_3` ADD `tema_81` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_80`;
# ALTER TABLE `disciplina_3` ADD `tema_82` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_81`;
# ALTER TABLE `disciplina_3` ADD `tema_83` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_82`;
# ALTER TABLE `disciplina_3` ADD `tema_84` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_83`;
# ALTER TABLE `disciplina_3` ADD `tema_85` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_84`;
# ALTER TABLE `disciplina_3` ADD `tema_86` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_85`;
# ALTER TABLE `disciplina_3` ADD `tema_87` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_86`;
# ALTER TABLE `disciplina_3` ADD `tema_88` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_87`;
# ALTER TABLE `disciplina_3` ADD `tema_89` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_88`;
# ALTER TABLE `disciplina_3` ADD `tema_90` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_89`;
# ALTER TABLE `disciplina_3` ADD `tema_91` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_90`;
# ALTER TABLE `disciplina_3` ADD `tema_92` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_91`;
# ALTER TABLE `disciplina_3` ADD `tema_93` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_92`;
# ALTER TABLE `disciplina_3` ADD `tema_94` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_93`;
# ALTER TABLE `disciplina_3` ADD `tema_95` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_94`;
# ALTER TABLE `disciplina_3` ADD `tema_96` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_95`;
# ALTER TABLE `disciplina_3` ADD `tema_97` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_96`;
# ALTER TABLE `disciplina_3` ADD `tema_98` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_97`;
# ALTER TABLE `disciplina_3` ADD `tema_99` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_98`;
# ALTER TABLE `disciplina_3` ADD `tema_100` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_99`;
# ALTER TABLE `disciplina_3` ADD `tema_101` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_100`;
# ALTER TABLE `disciplina_3` ADD `tema_102` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_101`;
# ALTER TABLE `disciplina_3` ADD `tema_103` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_102`;
# ALTER TABLE `disciplina_3` ADD `tema_104` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_103`;
# ALTER TABLE `disciplina_3` ADD `tema_105` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_104`;
# ALTER TABLE `disciplina_3` ADD `tema_106` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_105`;
# ALTER TABLE `disciplina_3` ADD `tema_107` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_106`;
# ALTER TABLE `disciplina_3` ADD `tema_108` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_107`;
# ALTER TABLE `disciplina_3` ADD `tema_109` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_108`;
# ALTER TABLE `disciplina_3` ADD `tema_110` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_109`;
# ALTER TABLE `disciplina_3` ADD `tema_111` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_110`;
# ALTER TABLE `disciplina_3` ADD `tema_112` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_111`;
# ALTER TABLE `disciplina_3` ADD `tema_113` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_112`;
# ALTER TABLE `disciplina_3` ADD `tema_114` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_113`;
# ALTER TABLE `disciplina_3` ADD `tema_115` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_114`;
# ALTER TABLE `disciplina_3` ADD `tema_116` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_115`;
# ALTER TABLE `disciplina_3` ADD `tema_117` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_116`;
# ALTER TABLE `disciplina_3` ADD `tema_118` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_117`;
# ALTER TABLE `disciplina_3` ADD `tema_119` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_118`;
# ALTER TABLE `disciplina_3` ADD `tema_120` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_119`;
# ALTER TABLE `disciplina_3` ADD `tema_121` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_120`;
# ALTER TABLE `disciplina_3` ADD `tema_122` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_121`;
# ALTER TABLE `disciplina_3` ADD `tema_123` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_122`;
# ALTER TABLE `disciplina_3` ADD `tema_124` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_123`;
# ALTER TABLE `disciplina_3` ADD `tema_125` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_124`;
# ALTER TABLE `disciplina_3` ADD `tema_126` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_125`;
# ALTER TABLE `disciplina_3` ADD `tema_127` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_126`;
# ALTER TABLE `disciplina_3` ADD `tema_128` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_127`;
# ALTER TABLE `disciplina_3` ADD `tema_129` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_128`;
# ALTER TABLE `disciplina_3` ADD `tema_130` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_129`;
# ALTER TABLE `disciplina_3` ADD `tema_131` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_130`;
# ALTER TABLE `disciplina_3` ADD `tema_132` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_131`;
# ALTER TABLE `disciplina_3` ADD `tema_133` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_132`;
# ALTER TABLE `disciplina_3` ADD `tema_134` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_133`;
# ALTER TABLE `disciplina_3` ADD `tema_135` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_134`;
# ALTER TABLE `disciplina_3` ADD `tema_136` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_135`;
# ALTER TABLE `disciplina_3` ADD `tema_137` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_136`;
# ALTER TABLE `disciplina_3` ADD `tema_138` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_137`;
# ALTER TABLE `disciplina_3` ADD `tema_139` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_138`;
# ALTER TABLE `disciplina_3` ADD `tema_140` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_139`;
# ALTER TABLE `disciplina_3` ADD `tema_141` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_140`;
# ALTER TABLE `disciplina_3` ADD `tema_142` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_141`;
# ALTER TABLE `disciplina_3` ADD `tema_143` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_142`;
# ALTER TABLE `disciplina_3` ADD `tema_144` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_143`;
# ALTER TABLE `disciplina_3` ADD `tema_145` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_144`;
# ALTER TABLE `disciplina_3` ADD `tema_146` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_145`;
# ALTER TABLE `disciplina_3` ADD `tema_147` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_146`;
# ALTER TABLE `disciplina_3` ADD `tema_148` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_147`;
# ALTER TABLE `disciplina_3` ADD `tema_149` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_148`;
# ALTER TABLE `disciplina_3` ADD `tema_150` VARCHAR(1) NULL DEFAULT NULL AFTER `tema_149`;
